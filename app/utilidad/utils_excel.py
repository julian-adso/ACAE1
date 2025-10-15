from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def exportar_asistencia_excel(datos, empleado="Empleado Desconocido", filename="reporte_asistencia.xlsx"):
    """
    datos: lista de diccionarios con llaves:
    ['fecha', 'ingreso', 'salida', 'estado', 'motivo']
    empleado: nombre del empleado
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    # ==========================
    # 🎯 ENCABEZADO PRINCIPAL
    # ==========================
    ws.merge_cells('A1:E1')
    titulo = ws['A1']
    titulo.value = f"Reporte de Asistencia - {empleado}"
    titulo.font = Font(bold=True, size=14, color="FFFFFF")
    titulo.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Subtítulo con fecha de generación
    ws.merge_cells('A2:E2')
    subtitulo = ws['A2']
    subtitulo.value = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    subtitulo.font = Font(italic=True, color="555555")
    subtitulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # ==========================
    # 🧾 ENCABEZADOS DE TABLA
    # ==========================
    headers = ["Fecha", "Ingreso", "Salida", "Estado", "Motivo"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ==========================
    # 🧱 ESTILOS BASE
    # ==========================
    thin_border = Border(
        left=Side(style='thin', color="BFBFBF"),
        right=Side(style='thin', color="BFBFBF"),
        top=Side(style='thin', color="BFBFBF"),
        bottom=Side(style='thin', color="BFBFBF")
    )

    colores_estado = {
        "Ausente": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),   # rojo claro
        "Presente": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"), # verde claro
        "Retardo": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # amarillo claro
        "Salida": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")    # gris claro
    }

    # ==========================
    # 📊 DATOS
    # ==========================
    for fila, item in enumerate(datos, start=4):
        estado = item.get('estado', '')
        ingreso_valor = item.get('ingreso', '')
        salida_valor = item.get('salida', '')

        # Fecha
        fecha_cell = ws.cell(row=fila, column=1, value=item.get('fecha', ''))
        fecha_cell.border = thin_border
        fecha_cell.alignment = Alignment(horizontal="center")

        # Ingreso (crear la celda primero)
        ingreso_cell = ws.cell(row=fila, column=2, value=ingreso_valor)
        ingreso_cell.border = thin_border
        ingreso_cell.alignment = Alignment(horizontal="center")

        # Salida (crear la celda primero)
        salida_cell = ws.cell(row=fila, column=3, value=salida_valor)
        salida_cell.border = thin_border
        salida_cell.alignment = Alignment(horizontal="center")

        # Estado
        estado_cell = ws.cell(row=fila, column=4, value=estado)
        estado_cell.border = thin_border
        estado_cell.alignment = Alignment(horizontal="center")
        if estado in colores_estado:
            estado_cell.fill = colores_estado[estado]

        # Motivo
        motivo_cell = ws.cell(row=fila, column=5, value=item.get('motivo', ''))
        motivo_cell.border = thin_border
        motivo_cell.alignment = Alignment(horizontal="center")

        # === Si está Ausente, poner "—" y aplicar fill gris claro ===
        if estado == "Ausente":
            ausente_fill = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")

            # Mostrar guion en ingreso/salida incluso si son "00:00" o vacíos
            if not ingreso_valor or ingreso_valor == "00:00":
                ingreso_cell.value = "—"
            if not salida_valor or salida_valor == "00:00":
                salida_cell.value = "—"

            ingreso_cell.fill = ausente_fill
            salida_cell.fill = ausente_fill


    # ==========================
    # 📏 AJUSTAR ANCHOS
    # ==========================
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # ==========================
    # 🔢 RESUMEN FINAL
    # ==========================
    total_filas = len(datos) + 5
    estados = [d['estado'] for d in datos]
    resumen = {
        "Presente": estados.count("Presente"),
        "Retardo": estados.count("Retardo"),
        "Ausente": estados.count("Ausente")
    }

    ws.merge_cells(f"A{total_filas}:E{total_filas}")
    resumen_titulo = ws[f"A{total_filas}"]
    resumen_titulo.value = "Resumen de Asistencia"
    resumen_titulo.font = Font(bold=True, size=12, color="FFFFFF")
    resumen_titulo.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    resumen_titulo.alignment = Alignment(horizontal="center")

    total_filas += 1
    for estado, cantidad in resumen.items():
        ws.merge_cells(f"A{total_filas}:D{total_filas}")
        ws[f"A{total_filas}"].value = estado
        ws[f"A{total_filas}"].font = Font(bold=True)
        ws[f"A{total_filas}"].alignment = Alignment(horizontal="right")
        ws[f"E{total_filas}"].value = cantidad
        ws[f"E{total_filas}"].alignment = Alignment(horizontal="center")
        total_filas += 1

    # ==========================
    # ✅ FILTROS Y GUARDADO
    # ==========================
    ws.auto_filter.ref = f"A3:E{len(datos) + 3}"
    wb.save(filename)
    return filename
